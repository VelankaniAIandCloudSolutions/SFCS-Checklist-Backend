import os
import time
import requests
import threading
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from pynput import keyboard

# Define the Excel file path
excel_file_path = '/home/velankani/Desktop/Work/SFCS/SFCS-Checklist-Backend/new_test.xlsx'  # Replace with the path to your Excel file

# Define the API endpoint URL
api_url = 'http://localhost:8000/api/v1/store/scan-code/'

# Global variable to track the last processed row number
last_processed_row = 1

# Function to send data to the API
def send_data_to_api(data):
    try:
        response = requests.post(api_url, json=data)
        if response.status_code == 200:
            print("Data sent to the API successfully!")
        else:
            print(f"Failed to send data to the API. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error sending data to the API: {str(e)}")

# Function to monitor the Excel file for changes
class ExcelFileHandler(FileSystemEventHandler):
    def on_modified(self, event):
        global last_processed_row
        
        # Trigger the API call when the Excel file is modified
        if event.src_path == excel_file_path:
            print("Excel file has been modified. Sending new data to API...")
            
            # Read the Excel file and send data to the API for newly added rows
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # Iterate through rows starting from the last processed row
            for row_num, row in enumerate(worksheet.iter_rows(min_row=last_processed_row, min_col=1, max_col=1, values_only=True), start=last_processed_row):
                cell_value = row[0]
                if cell_value:
                    data = {'value': cell_value}
                    send_data_to_api(data)
            
            # Update the last processed row number
            last_processed_row = row_num + 1
            
            # Close the workbook
            workbook.close()

# Function to monitor Enter key press
# (Assuming you want the same action as the file modification)
def monitor_enter_key():
    def on_key_release(key):
        global last_processed_row
        
        if key == keyboard.Key.enter:
            print("Enter key pressed. Sending new data to API...")
            
            # Read the Excel file and send data to the API for newly added rows
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # Iterate through rows starting from the last processed row
            for row_num, row in enumerate(worksheet.iter_rows(min_row=last_processed_row, min_col=1, max_col=1, values_only=True), start=last_processed_row):
                cell_value = row[0]
                if cell_value:
                    data = {'value': cell_value}
                    send_data_to_api(data)
            
            # Update the last processed row number
            last_processed_row = row_num + 1
            
            # Close the workbook
            workbook.close()
    
    # Start the keyboard listener
    with keyboard.Listener(on_release=on_key_release) as listener:
        listener.join()

if __name__ == "__main__":
    # Start the Excel file modification monitoring
    observer = Observer()
    observer.schedule(ExcelFileHandler(), path=os.path.dirname(excel_file_path), recursive=False)
    observer.start()

    # Start the Enter key monitoring in a separate thread
    enter_key_thread = threading.Thread(target=monitor_enter_key)
    enter_key_thread.daemon = True
    enter_key_thread.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()

    observer.join()
